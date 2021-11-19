'''
Created on Nov 18, 2021

@author: nifty
'''

import json
import traceback
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import sys


import base64
from Helpers import DataPull

absolutepath = os.path.abspath(__file__)
fileDirectory = os.path.dirname(absolutepath)
parentDirectory = os.path.dirname(fileDirectory)
newPath = os.path.join(parentDirectory, 'STConnectorCLI') 
sitename = ""
apikey = ""
def gen_file(wanted_fields):
    
    #Declarations and assignment
    job_records = get_jobs()
    wb = Workbook()
    dest_file = newPath+"\\"+file_n+".xlsx"
    report_worksheet = wb.active
    report_worksheet.title = "Report"
    r = 2
     
    data_worksheet = wb.create_sheet(title = "Data")
   
    try:
        for sf in wanted_fields:
            data_worksheet.cell(row=1, column=wanted_fields.index(sf)+1, value = sf)
        for job in job_records:
            lrr = []
            c = 1
            for sf in wanted_fields:
                res = ""
                if "_" in sf:
                    try:
                        sf_l = sf.split("_")
                        if sf_l[0] == "custom":
                            for l in job['customFieldValues']:
                                if l['label'] == sf_l[1]:
                                    res = l['value']
                        else:
                            res = job[sf_l[0]][sf_l[1]]
                    except:
                        pass                        
                else:
                    res = job[sf]
                data_worksheet.cell(row=r, column=c, value=str(res))
                lrr.append(str(res))
                c += 1
            print(lrr)
            r += 1
    except:
        pass
    wb.save(filename = dest_file)
              
      

def load_job_fields(fields_list):
    try:
        mn_fields_list = fields_list + open(newPath+"\\Helpers\\fields.txt","r").read().split("\n")
        print("\n".join(mn_fields_list))
    except:
        pass
def get_jobs():
    try:
        
        x = 1
        job_records = []
        while x<100:
            nx = DataPull.pull_jobs(x, base64_bytes)
            print(str(nx)[0:int(len(str(nx))/4)])
            job_records += nx['data']
            if len(nx['data']) == 0:
                x = 1001
                continue
            x += 1
        return job_records
    except:
        pass
def get_customs():
    try:
        
        data = DataPull.pull_customs(base64_bytes)
        ld = []
        for l in data['data']:
            ld.append("custom_"+l['label'])
        load_job_fields(ld)
    except:
        pass
message_bytes = str(input("Enter sitename:apikey:\n")).encode('ascii')
base64_bytes = base64.urlsafe_b64encode(message_bytes)
print(base64_bytes)
file_n = "default"
while True:
    choice = input("a) Get Fields (inc. Custm Fields) b) Generate Excel File c) Exit:\n")
    if choice == "a":
        get_customs()
    elif choice == "b":
        file_n = input("Enter File Export Name w/o .extension:\n")
        if file_n == "":
            file_n = "default"
        gen_file(input("Enter fields separated by space:\n").split(" "))
    elif choice == "c":
        break

